#!/usr/bin/env python3
"""Generate a job-search tracker.

Writes an .xlsx workbook (Applications, Contacts, Source summary) and CSV
exports of each tab, so the same schema can be imported into Google Sheets,
Notion or Airtable.

Usage:
    python3 make_tracker.py                     # -> ./job-search-tracker.xlsx + CSVs
    python3 make_tracker.py --out ~/job-search  # choose a directory
    python3 make_tracker.py --csv-only          # skip the xlsx (no openpyxl needed)

No personal data is embedded. One example row is included and marked EXAMPLE -
delete it once real roles are added.
"""

import argparse
import csv
import os
import sys

APPLICATIONS = [
    "Rank", "Date found", "Source", "Company", "Role title", "Band", "Depth",
    "Via", "Consent",
    "Location / mode", "JD link", "Posted", "Salary band",
    "Fit", "Probability", "Priority",
    "Why", "Watch", "Route", "Advocate",
    "Status", "Date applied", "Next action", "Due", "CV used", "Notes",
]

# Where a listing hides the employer, Company reads
# "[EMPLOYER NOT DISCLOSED]" and Via names the agency. Never guess the employer -
# two hidden listings are often the same underlying role, and a guess makes that
# undetectable.
CONSENT = [
    "Direct - n/a",
    "Asked, awaiting client name",
    "Consent given",
    "Consent withheld",
    "SUBMITTED WITHOUT CONSENT",
]

HISTORY = [
    "Company", "Role", "Level", "Date", "Stage reached", "Outcome", "Feedback",
    "Notes",
]

CONTACTS = [
    "Name", "Company", "Relationship", "How they can help", "Asked on",
    "Last contact", "Next follow-up", "Status", "Notes",
]

SOURCES = ["Source", "Roles found", "Applications", "Recruiter calls",
           "In process", "Offers"]

# Rejected (they declined you) and Passed (you declined them, or closed a dead
# application at day 30) are deliberately distinct - the Passed rows are where
# your screening filters come from.
STATUSES = [
    "Alert only", "Not started", "Researching", "Applied", "Recruiter call",
    "In process", "Onsite/Final", "Offer", "Rejected", "Passed",
]

DEPTHS = ["High", "Medium", "Low"]

BANDS = ["Mid", "Senior", "Lead", "Staff", "Principal", "Director", "VP", "Other"]

DEFAULT_SOURCES = [
    "LinkedIn", "Local job board", "Greenhouse (my.greenhouse.io)",
    "Company board", "Referral", "Recruiter", "Other",
]

EXAMPLE_ROW = [
    "1", "2026-08-16", "LinkedIn", "EXAMPLE - delete this row",
    "Senior Product Manager", "Senior", "High", "", "Direct - n/a",
    "Dublin / hybrid 2d", "https://",
    "req 12345, posted 6 days ago", "",
    "8", "6", "",
    "Core of the role is the thing I am best at, not a bolt-on",
    "They name a tool I have never used - do not blur it with the adjacent one",
    "Direct application, no route in yet", "",
    "Not started", "", "Read full JD and re-score", "2026-08-18", "", "",
]


def write_csvs(out_dir):
    paths = []
    for name, header, rows in (
        ("applications", APPLICATIONS, [EXAMPLE_ROW]),
        ("contacts", CONTACTS, []),
        ("interview-history", HISTORY, []),
        ("source-summary", SOURCES, [[s, "", "", "", "", ""] for s in DEFAULT_SOURCES]),
    ):
        path = os.path.join(out_dir, f"job-search-tracker-{name}.csv")
        with open(path, "w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            writer.writerow(header)
            writer.writerows(rows)
        paths.append(path)
    return paths


def write_xlsx(out_dir):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        print("openpyxl not installed - run 'pip install openpyxl' or use --csv-only")
        return None

    wb = Workbook()
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="2F4858")

    def style_header(ws, header, widths):
        ws.append(header)
        for i, _ in enumerate(header, start=1):
            cell = ws.cell(row=1, column=i)
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(i)].width = widths.get(i, 16)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}1"

    ws = wb.active
    ws.title = "Applications"
    style_header(ws, APPLICATIONS, {
        4: 22, 5: 28, 8: 20, 9: 26, 10: 22, 11: 30, 12: 24,
        17: 40, 18: 40, 19: 30, 23: 30, 26: 34})
    ws.append(EXAMPLE_ROW)

    # Column letters are derived from APPLICATIONS, never hard-coded - inserting
    # a column silently repointed every one of these the last time they were.
    col = {name: get_column_letter(i)
           for i, name in enumerate(APPLICATIONS, start=1)}

    status_dv = DataValidation(type="list", formula1='"%s"' % ",".join(STATUSES), allow_blank=True)
    depth_dv = DataValidation(type="list", formula1='"%s"' % ",".join(DEPTHS), allow_blank=True)
    band_dv = DataValidation(type="list", formula1='"%s"' % ",".join(BANDS), allow_blank=True)
    source_dv = DataValidation(type="list", formula1='"%s"' % ",".join(DEFAULT_SOURCES), allow_blank=True)
    consent_dv = DataValidation(type="list", formula1='"%s"' % ",".join(CONSENT), allow_blank=True)
    score_dv = DataValidation(type="whole", operator="between", formula1=1, formula2=10, allow_blank=True)
    for dv in (status_dv, band_dv, source_dv, score_dv, depth_dv, consent_dv):
        ws.add_data_validation(dv)
    status_dv.add(f"{col['Status']}2:{col['Status']}500")
    band_dv.add(f"{col['Band']}2:{col['Band']}500")
    depth_dv.add(f"{col['Depth']}2:{col['Depth']}500")
    source_dv.add(f"{col['Source']}2:{col['Source']}500")
    consent_dv.add(f"{col['Consent']}2:{col['Consent']}500")
    score_dv.add(f"{col['Fit']}2:{col['Fit']}500")
    score_dv.add(f"{col['Probability']}2:{col['Probability']}500")

    # Priority = Fit and Probability weighted equally. Overwrite by hand when
    # comp or title justifies an adjustment, and say so in the Why column.
    f, pr = col["Fit"], col["Probability"]
    for r in range(2, 501):
        ws.cell(row=r, column=APPLICATIONS.index("Priority") + 1,
                value=f'=IF(COUNT({f}{r}:{pr}{r})=2,AVERAGE({f}{r}:{pr}{r}),"")')

    ws_c = wb.create_sheet("Contacts")
    style_header(ws_c, CONTACTS, {4: 30, 9: 34})

    ws_h = wb.create_sheet("Interview history")
    style_header(ws_h, HISTORY, {5: 22, 7: 40, 8: 34})

    ws_s = wb.create_sheet("Source summary")
    style_header(ws_s, SOURCES, {1: 28})
    for i, src in enumerate(DEFAULT_SOURCES, start=2):
        ws_s.cell(row=i, column=1, value=src)
        src_c, app_c, st_c = col["Source"], col["Date applied"], col["Status"]
        ws_s.cell(row=i, column=2, value=(
            f'=COUNTIF(Applications!${src_c}$2:${src_c}$500,$A{i})'))
        ws_s.cell(row=i, column=3, value=(
            f'=COUNTIFS(Applications!${src_c}$2:${src_c}$500,$A{i},'
            f'Applications!${app_c}$2:${app_c}$500,"<>")'
        ))
        for c, stage in ((4, "Recruiter call"), (5, "In process"), (6, "Offer")):
            ws_s.cell(row=i, column=c, value=(
                f'=COUNTIFS(Applications!${src_c}$2:${src_c}$500,$A{i},'
                f'Applications!${st_c}$2:${st_c}$500,"{stage}")'
            ))
    note_row = len(DEFAULT_SOURCES) + 3
    ws_s.cell(row=note_row, column=1,
              value="Applications counts rows with a 'Date applied'. Stage columns "
                    "count current status only - a role now at Offer no longer "
                    "counts as a Recruiter call. Compare warm routes (Advocate "
                    "filled) against cold ones at least monthly.")

    path = os.path.join(out_dir, "job-search-tracker.xlsx")
    wb.save(path)
    return path


def main():
    ap = argparse.ArgumentParser(description="Generate a job-search tracker.")
    ap.add_argument("--out", default=".", help="output directory (default: current)")
    ap.add_argument("--csv-only", action="store_true", help="skip the xlsx")
    args = ap.parse_args()

    out_dir = os.path.expanduser(args.out)
    os.makedirs(out_dir, exist_ok=True)

    written = write_csvs(out_dir)
    if not args.csv_only:
        xlsx = write_xlsx(out_dir)
        if xlsx:
            written.insert(0, xlsx)

    print("Wrote:")
    for p in written:
        print("  " + p)
    print("\nGoogle Sheets: File > Import > Upload the applications CSV, then repeat "
          "for each of the other CSVs as new tabs.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
